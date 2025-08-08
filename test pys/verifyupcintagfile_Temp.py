import time
import shutil
import os
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font


def upcremovalProcess():
    pwd = os.getcwd()
    upcFolderPath = os.path.join(pwd,"Test Data", "UPC")
    tagFolderpath = os.path.join(pwd,"Test Data", "Tag")

    arrUPCNotIn = []
    tagUPC = []


    tagFileCount = 0
    for tagfile in os.listdir(tagFolderpath):
        orgtagfile_Path = os.path.join(tagFolderpath, tagfile)
        extension = os.path.splitext(tagfile)[1]

        if not extension.lower() in [".xls", ".xlsx", ".xlsm", ".xlsb", ".csv", ".pdf", ".jpg", ".png", ".doc", ".docx", ".docm", ".bmp", ".msg", ".ppt", ".pptx", ".pptm"]:
            tagFileCount = tagFileCount + 1

            with open(orgtagfile_Path, 'r+') as sfile:
                data = sfile.readlines()

            for line in data:
                start = 45  # upc 45 - 57 (13 digits)
                end = 57
                tagupcVal = line[start:end]
                if tagupcVal.strip():
                    tagUPC.append(tagupcVal)



    for entry in os.listdir(upcFolderPath):
            full_path = os.path.join(upcFolderPath, entry)
            if os.path.isfile(full_path):
                extension = os.path.splitext(full_path)[1]
                if extension == ".xlsx":
                    upcFilePath = full_path


    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    workbook = openpyxl.load_workbook(upcFilePath)
    for eachSheet in workbook.worksheets:
        if eachSheet.sheet_state == "visible":
            # sheet = workbook.worksheets[i]
            for column in eachSheet.iter_cols():
                column_name = column[0].value
                if column_name == "UPC":
                    for i, cell in enumerate(column):
                        if i != 0:
                            row = cell.row
                            if eachSheet.row_dimensions[row].hidden:
                                continue
                            upcstr = str(cell.value)
                            upcstr = upcstr.replace("-", "")
                            if upcstr and upcstr != 'None' :
                                upcFound = False
                                for arrtagUPCval in tagUPC:
                                    if int(arrtagUPCval) == int(upcstr):
                                        upcFound = True

                                if upcFound != True:
                                    arrUPCNotIn.append(upcstr) 
                                    cell.fill = red_fill


    workbook.save(upcFilePath)

          
     

    for UPCNotIn in arrUPCNotIn:
        print(UPCNotIn)
                           


if __name__ == "__main__":
     upcremovalProcess()








